import keycloak
import openpyxl
import subprocess

keycloak_admin = keycloak.KeycloakAdmin(server_url="https://tigeranettrainer.com/auth/",
                               username='anetAdmin',
                               password='secret',
                               realm_name="master",
                               verify=False)

keycloak_admin.realm_name = "ANET-Realm"

book = openpyxl.load_workbook(filename = 'users.xlsx')

sheet = book.active

for sheet in book:
    print("> sheet ", sheet.title)
    for row in sheet.iter_rows(min_row=2, min_col=2, max_row=50, max_col=2):
        for cell in row:
            if cell.value and not sheet.cell(column=5, row=cell.row).value:
                username = cell.value.lower()
                password = subprocess.check_output(['pwgen', '10', '1']).decode("utf-8").replace('\n','')
                try:
                    user_id = keycloak_admin.get_user_id(username)
                    
                    if user_id is None:
                        new_user = keycloak_admin.create_user({"username": username,
                                        "enabled": True,
                                        "realmRoles": ["user_default", ]})
                        user_id = keycloak_admin.get_user_id(username)
                        print(" creating -> ", username, " ", password)
                    else:
                        print(" updating -> ", username, " ", password)

                    keycloak_admin.update_user(user_id=user_id,
                                    payload={"credentials": [{"value": password,"type": "password",}]})
                    
                    sheet.cell(column=2, row=cell.row).value=username
                    sheet.cell(column=5, row=cell.row).value=password
                    book.save('users.imported.xlsx')
                except:
                    logging.error(traceback.format_exc())
